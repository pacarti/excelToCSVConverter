import json, os, openpyxl, csv
from pathlib import Path
from openpyxl.utils import get_column_letter

os.chdir(os.path.dirname(os.path.abspath(__file__)))


# Skip non-xlsx files, load the workbook object.
for file in os.listdir(os.getcwd()):
    if file.endswith('.xlsx'):
        # print(file)
        wb = openpyxl.load_workbook(file)
        # print(wb.sheetnames)
    else:
        continue

    # Loop through every sheet in the workbook.    
    for sheetName in wb.sheetnames:
        
        sheet = wb[sheetName]

        # Create the CSV filename from the Excel filename and sheet title.
        pfile = Path(file)
        # print(pfile.stem + '_' + sheetName + '.csv')

        csvFile = open(pfile.stem + '_' + sheetName + '.csv', 'w')

        # Create the csv.writer object for this CSV file.
        outputWriter = csv.writer(csvFile)

        # Loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row + 1):
            rowData = [] # append each cell to this list      
            # Loop through each cell in the row.
            for colNum in range(1, sheet.max_column + 1):
                # Append each cell's data to rowData.
                rowData.append(sheet[get_column_letter(colNum) + str(rowNum)].value)
            # Write the rowData list to the CSV file.
            print(rowData)
csvFile.close()